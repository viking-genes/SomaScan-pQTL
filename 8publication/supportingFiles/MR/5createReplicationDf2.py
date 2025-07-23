import pandas as pd
import os

# --- Setup ---
projectRoot = './prj_190_viking_somalogic/'

# --- Define directories and file paths ---
mrResultDir = os.path.join(projectRoot, 'Scripts/Publication/supportingFiles/replication/MR/4resultsMR')
harmonisedPreMRDir = os.path.join(mrResultDir, 'harmonisedPreMRData')

replicationInputPath = os.path.join(projectRoot, 'Scripts/Publication/supportingFiles/replication/MR/3replicationDf.xlsx')
replicationOutputPath = os.path.join(projectRoot, 'Scripts/Publication/supportingFiles/replication/MR/5replicationDf.xlsx')

vikingMRPath = os.path.join(projectRoot, 'Scripts/Publication/supplementaryMR.xlsx')


df = pd.read_excel(df)
VIKINGMRdf = pd.read_excel(VIKINGMRdf)

def createMRresultDf(MRresultDir, VIKINGMRdf):

    #Calculate F-statistic for Wald Ratio MR instrument strength estimation
    #Formula to calculate the F-statistic for a single SNP https://academic.oup.com/ije/article/40/3/755/745918?login=true
    #F=VarianceExplained(Nsamples-2)/(1-VarianceExplained)
    def calculateFStatistic(df, singleSNP=False):
        
        #Open the file with the pQTL data for GWAS summary statistics
        exposure = df.loc[0, 'MR Exposure']
        outcomeID = df.loc[0, 'Study ID']
        assay = df.loc[0, 'assay']
        filename = f'MRHarmonised_{exposure}_{assay}_{outcomeID}.tsv'
        preMRdf = pd.read_csv(os.path.join(preMRdir, filename), sep='\t')
        
        #Process single SNP data
        if df.loc[0, 'Replication MR Number of SNPs'] == 1:
            #Calculate the variance explained by the exposure
            exposureVarianceExplained = 2 * preMRdf.loc[0, 'eaf.exposure'] * (1 - preMRdf.loc[0, 'eaf.exposure']) * preMRdf.loc[0, 'beta.exposure']**2  
            exposureSampleSize = preMRdf.loc[0, 'samplesize.exposure']

            fStatistic = exposureVarianceExplained * (exposureSampleSize - 2) / (1 - exposureVarianceExplained)

            fStatistic = round(fStatistic, 2)
        
        #Process multiple SNP data
        else:

            FStatisticList = []

            for i in range(len(preMRdf)):
                #Calculate the variance explained by the exposure
                exposureVarianceExplained = 2 * preMRdf.loc[i, 'eaf.exposure'] * (1 - preMRdf.loc[i, 'eaf.exposure']) * preMRdf.loc[i, 'beta.exposure']**2  
                exposureSampleSize = preMRdf.loc[i, 'samplesize.exposure']

                fStatistic = exposureVarianceExplained * (exposureSampleSize - 2) / (1 - exposureVarianceExplained)

                FStatisticList.append(round(fStatistic, 2))
            
            #Return a value of average [min, max]
            fStatistic = f'{round(sum(FStatisticList) / len(FStatisticList), 2)} [{min(FStatisticList)}-{max(FStatisticList)}]'

        return fStatistic


    result = pd.DataFrame({
        'MR Exposure': pd.Series(dtype='str'),
        'assay': pd.Series(dtype='str'),
        'Study ID': pd.Series(dtype='str'),
        'beta': pd.Series(dtype='float'),
        'se': pd.Series(dtype='float'),
        'p': pd.Series(dtype='float'),
        'Discovery MR Outcome': pd.Series(dtype='str'),
        'Discovery MR Effect Size (beta)': pd.Series(dtype='float'),
        'Discovery MR p-value': pd.Series(dtype='float'),
        'Discovery MR Outcome Study Publication Year': pd.Series(dtype='str'),
        'Discovery MR Outcome Study Sample Size': pd.Series(dtype='str'),
        'Discovery MR Outcome Study pmid': pd.Series(dtype='str'),
        'Replication MR Number of SNPs': pd.Series(dtype='int'),
        'Replication MR Method': pd.Series(dtype='str'),
        'Replication MR Instrumental variable F-statistic': pd.Series(dtype='object')  # F-statistic can be a string when in the [min-max] format
    })

    for filename in os.listdir(MRresultDir):
        if filename.endswith('.tsv'):
            df = pd.read_csv(os.path.join(MRresultDir, filename), sep='\t')

            geneName = filename.split('_')[1]
            assay = filename.split('_')[2]
            outcome = ('_').join(filename.split('_')[3:])[:-4]

            #Find the row in the VIKING Discovery MR data that matches the MR replication result
            VIKINGdiscoveryRow = VIKINGMRdf[(VIKINGMRdf['MR Exposure'] == geneName) & (VIKINGMRdf['Study ID'] == outcome)]

            # Create a new row as a DataFrame
            newRow = pd.DataFrame({
                'MR Exposure': [geneName],
                'assay': [assay],
                'Study ID': [outcome],
                'beta': [df.loc[0, 'b']],
                'se': [df.loc[0, 'se']],
                'p': [df.loc[0, 'pval']],
                'Discovery MR Outcome': [df.loc[0, 'outcome']][0].split(' || ')[0],
                'Discovery MR Effect Size (beta)': [VIKINGdiscoveryRow['MR Effect size (beta)'].values[0]],
                'Discovery MR p-value': [VIKINGdiscoveryRow['MR p-value'].values[0]],
                'Discovery MR Outcome Study Publication Year': [df.loc[0, 'year']],
                'Discovery MR Outcome Study Sample Size': [df.loc[0, 'sample_size']],
                'Discovery MR Outcome Study pmid': [df.loc[0, 'pmid']],
                'Replication MR Number of SNPs': [df.loc[0, 'nsnp.x']],
                'Replication MR Method': [df.loc[0, 'method']]
            })

            # Calculate the F-statistic for the MR instrument strength estimation
            fStatistic = calculateFStatistic(newRow)

            newRow['Replication MR Instrumental variable F-statistic'] = fStatistic

            # Append the new row to the result DataFrame using concat
            result = pd.concat([result, newRow], ignore_index=True)
            
    return result

def addAllMRResults(df, MRResultdf):

    #Get the MR results that have not been processed yet
    MRResultdf = MRResultdf[MRResultdf['processed'] == False]
    MRResultdf = MRResultdf.drop(columns='processed')

    #Rename the columns to match the replication DataFrame
    MRResultdf = MRResultdf.rename(columns={'MR Exposure' : 'Replication MR Exposure', 'Study ID' : 'Discovery MR Study ID', 'beta' : 'Replication MR Stage 2 effect size (beta)', 'se' : 'Replication MR Stage 2 beta standard error', 'p' : 'Replication MR Stage 2 p-value', 'Replication MR Number of SNPs' : 'Replication MR Stage 2 Number of SNPs', 'Replication MR Method' : 'Replication MR Stage 2 Method', 'Replication MR Instrumental variable F-statistic' : 'Replication MR Stage 2 Instrumental variable F-statistic'})

    #Combine the Replication MR Exposure and assay columns into one
    MRResultdf['Replication MR Exposure'] = MRResultdf['Replication MR Exposure'] + '_' + MRResultdf['assay']
    MRResultdf = MRResultdf.drop(columns='assay')

    #Add each MR result to the DataFrame as a new row
    for i in MRResultdf.index:

        newRow = MRResultdf.loc[i].copy()

        #Round the effect size and p-value 
        newRow['Replication MR Stage 2 effect size (beta)'] = round(newRow['Replication MR Stage 2 effect size (beta)'], 5)
        newRow['Replication MR Stage 2 beta standard error'] = round(newRow['Replication MR Stage 2 beta standard error'], 5)
        newRow['Replication MR Stage 2 p-value'] = float(f"{newRow['Replication MR Stage 2 p-value']:.5g}")

        #Add the new row to the DataFrame
        df = pd.concat([df, pd.DataFrame([newRow])], ignore_index=True)

    return df

#Removes data for MR Stage 2 results from the dataframe if Stage 1 was significant
def correctBonferroni(df, threshold):
    
    df = df.reset_index(drop=True)

    columnsToRemove = ['Replication MR Stage 2 effect size (beta)', 'Replication MR Stage 2 beta standard error', 'Replication MR Stage 2 p-value', 'Replication MR Stage 2 Number of SNPs', 'Replication MR Stage 2 Method', 'Replication MR Stage 2 Instrumental variable F-statistic']

    for i in range(len(df)):
        if df.loc[i, 'Replication MR Stage 1 p-value'] < threshold:
            for column in columnsToRemove:
                df.at[i, column] = 'NA'

    return df

def sortByPvalue(df, threshold):
    
    #Split the dataframe into two, based on the Bonferroni threshold
    df1 = df[df['Replication MR Stage 1 p-value'] < bonferroniThreshold]
    df2 = df.drop(df1.index)

    #Sort df1 by MR p-value
    df1 = df1.sort_values(by='Replication MR Stage 1 p-value')

    pValues = df2['Replication MR Stage 2 p-value'].tolist()
    pValues = [str(x).split('|') for x in pValues]
    pValues = [[float(y) for y in x] for x in pValues]
    pValues = [min(x) for x in pValues]

    #Sort df2 by MR p-value
    df2 = df2.assign(pValue=pValues)
    df2 = df2.sort_values(by='pValue')
    df2 = df2.drop(columns='pValue')

    #Concatenate the two dataframes
    df = pd.concat([df1, df2], ignore_index=True)

    #Draw a line after the last significant MR result
    df = df.reset_index(drop=True)

    return df

#Add a column with True/False if the MR result is significant in either stage
def addSignificantColumn(df, threshold):

    significantList = []

    for i in range(len(df)):
        
        geneName = df.loc[i, 'Replication MR Exposure'].split('_')[0]
        outcome = df.loc[i, 'Replication MR Stage 1 Outcome']

        currentValue = 'No'

        pValueList = str(df.loc[i, 'Replication MR Stage 1 p-value']).split('|') + str(df.loc[i, 'Replication MR Stage 2 p-value']).split('|')

        #Remove any NA values
        pValueList = [x for x in pValueList if x != 'NA']

        #Convert to float
        pValueList = [float(x) for x in pValueList]

        for pValue in pValueList:
            if pValue < threshold:
                currentValue = 'Yes'
                break
        
        significantList.append(currentValue)

    df['MR Significant result'] = significantList

    return df

#Calculate the Bonferroni threshold for the number of tests performed in stage 1 replication and total possible number of tests in stage 2 MR
def calculateBonferroniThreshold(df):

    #Count the total number of tests that can be performed in stage 2
    outcomesTested = df['Replication MR Stage 2 p-value'].tolist()
    outcomesTested = [str(x).split('|') for x in outcomesTested]
    #Flatten the list
    outcomesTested = [item for sublist in outcomesTested for item in sublist]

    #Add the stage 1 p-values to the list excluding NA values
    outcomesTested = outcomesTested + df['Replication MR Stage 1 p-value'].tolist()
    outcomesTested = [x for x in outcomesTested if x != 'NA' and x == x]
    
    #bonferroniThreshold = 0.000472
    result = 0.05 / len(outcomesTested)

    return result


MRresultDf = createMRresultDf(MRresultDir, VIKINGMRdf)

#Add a new column to track which MR results have been processed
MRresultDf['processed'] = False

#Create new columns to be added
df['Replication MR Stage 2 effect size (beta)'] = ''
df['Replication MR Stage 2 beta standard error'] = ''
df['Replication MR Stage 2 p-value'] = ''
df['Replication MR Stage 2 Number of SNPs'] = ''
df['Replication MR Stage 2 Method'] = ''
df['Replication MR Stage 2 Instrumental variable F-statistic'] = ''

#Merge the MR result data with the replication data
for i in range(len(df)):

    replicationExposure = df.loc[i, 'Replication MR Exposure'].split('_')[0]
    replicationAssay = df.loc[i, 'Replication MR Exposure'].split('_')[1]
    replicationStudyListStage2 = df.loc[i, 'Discovery MR Study ID'].split('|')

    #Find the MR result data for the replication exposure
    MRresult = MRresultDf[(MRresultDf['MR Exposure'] == replicationExposure) & (MRresultDf['assay'] == replicationAssay)]

    dataDict = {'Replication MR Stage 2 effect size (beta)' : [], 'Replication MR Stage 2 beta standard error' : [], 'Replication MR Stage 2 p-value' : [], 'Replication MR Stage 2 Number of SNPs' : [], 'Replication MR Stage 2 Method' : [], 'Replication MR Stage 2 Instrumental variable F-statistic' : []}

    for studyID in replicationStudyListStage2:
        #Get the appropriate row from the MR result data
        replicationRow = MRresult[MRresult['Study ID'] == studyID]

        #Mark the selected rows as processed
        MRresultDf.loc[replicationRow.index, 'processed'] = True

        #Add the MR result data to the replication dataDict
        dataDict['Replication MR Stage 2 effect size (beta)'].append(replicationRow['beta'].values[0])
        dataDict['Replication MR Stage 2 beta standard error'].append(replicationRow['se'].values[0])
        dataDict['Replication MR Stage 2 p-value'].append(replicationRow['p'].values[0])
        dataDict['Replication MR Stage 2 Number of SNPs'].append(replicationRow['Replication MR Number of SNPs'].values[0])
        dataDict['Replication MR Stage 2 Method'].append(replicationRow['Replication MR Method'].values[0])
        dataDict['Replication MR Stage 2 Instrumental variable F-statistic'].append(replicationRow['Replication MR Instrumental variable F-statistic'].values[0])

    #Round the effect size and p-value 
    dataDict['Replication MR Stage 2 effect size (beta)'] = [round(x, 5) for x in dataDict['Replication MR Stage 2 effect size (beta)']]
    dataDict['Replication MR Stage 2 beta standard error'] = [round(x, 5) for x in dataDict['Replication MR Stage 2 beta standard error']]
    dataDict['Replication MR Stage 2 p-value'] = [float(f"{x:.5g}") for x in dataDict['Replication MR Stage 2 p-value']]

    #Add the dataDict to the replication DataFrame
    df.loc[i, 'Replication MR Stage 2 effect size (beta)'] = '|'.join(map(str, dataDict['Replication MR Stage 2 effect size (beta)']))
    df.loc[i, 'Replication MR Stage 2 beta standard error'] = '|'.join(map(str, dataDict['Replication MR Stage 2 beta standard error']))
    df.loc[i, 'Replication MR Stage 2 p-value'] = '|'.join(map(str, dataDict['Replication MR Stage 2 p-value']))
    df.loc[i, 'Replication MR Stage 2 Number of SNPs'] = '|'.join(map(str, dataDict['Replication MR Stage 2 Number of SNPs']))
    df.loc[i, 'Replication MR Stage 2 Method'] = '|'.join(map(str, dataDict['Replication MR Stage 2 Method']))
    df.loc[i, 'Replication MR Stage 2 Instrumental variable F-statistic'] = '|'.join(map(str, dataDict['Replication MR Stage 2 Instrumental variable F-statistic']))


#Add MR results for Stage 2
df = addAllMRResults(df, MRresultDf)

#Calculate the Bonferroni threshold
bonferroniThreshold = calculateBonferroniThreshold(df)

#Assign NA for Stage 2 rows that that replicate in Stage 1
df = correctBonferroni(df, bonferroniThreshold)

#Sort the DataFrame by MR p-value
df = sortByPvalue(df, bonferroniThreshold)

#Fill the dataframe with NA for any missing values
df = df.fillna('NA')

#Reorder the columns, adding the new stage 2 columns after the stage 1 columns
df = df[['Replication MR Exposure', 'Replication MR Stage 1 Outcome', 'Replication MR Stage 1 Effect size (beta)', 'Replication MR Stage 1 Standard Error of beta', 'Replication MR Stage 1 p-value', 'Replication MR Number of SNPs', 'Replication MR Stage 2 effect size (beta)', 'Replication MR Stage 2 beta standard error', 'Replication MR Stage 2 p-value', 'Replication MR Stage 2 Number of SNPs', 'Discovery MR Outcome', 'Discovery MR Effect Size (beta)', 'Discovery MR p-value', 'Replication MR Stage 1 Study ID', 'Replication MR Stage 1 Method', 'Replication MR Stage 1 Instrumental variable F-statistic', 'Replication Stage 1 Outcome Study Author', 'Replication Stage 1 Outcome Study Publication Year', 'Replication Stage 1 Outcome Study Sample Size', 'Replication Stage 1 Outcome Study pmid', 'Replication MR Stage 2 Method', 'Replication MR Stage 2 Instrumental variable F-statistic', 'Discovery MR Study ID', 'Discovery MR Outcome Study Publication Year', 'Discovery MR Outcome Study Sample Size', 'Discovery MR Outcome Study pmid']]

#Add a column to indicate if the MR result is significant in either stage
df = addSignificantColumn(df, bonferroniThreshold)

#Save the replication DataFrame
df.to_excel(outputPath, index=False)

#Apply the thick border to the replication DataFrame based on the Bonferroni threshold
def applyBonferroniThickBorder(filepath, sheet_name, df, column_name, threshold):
    
    import openpyxl
    from openpyxl.styles import Border, Side

    # Load the Excel file
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheet_name]

    # Define the thick border style
    thick_border = Border(
        top=Side(style='double')
    )

    # Get the index of the column to apply the threshold on
    col_index = df.columns.get_loc(column_name) + 1

    # Iterate through the DataFrame and check the condition
    for row in range(2, len(df) + 2):  # Start from row 2 (excluding the header)
        #Skip rows where the value is NA
        if df.loc[row - 2, column_name] == 'NA':
            continue
        
        pValue = str(df.loc[row - 2, column_name]).split('|')
        pValue = [float(x) for x in pValue]
        pValue = min(pValue)

        if pValue >= threshold:
            # Apply the thick border to the entire row
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thick_border
        
            break

    # Save the updated workbook
    workbook.save(filepath)
    print(f"Thick border applied to rows where {column_name} >= {threshold}.")

# Apply the thick border to the replication DataFrame
applyBonferroniThickBorder(outputPath, 'Sheet1', df, 'Replication MR Stage 2 p-value', bonferroniThreshold)

